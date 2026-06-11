from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

REQUIRED_COLUMNS = [
    "Proyecto",
    "Consultor",
    "Analista responsable",
    "Proveedor",
    "Perfil",
    "Costo Mensual [USD]",
    "Costo Mensual [PEN]",
    "Duración",
    "Costo Total [USD]",
    "Costo Total [PEN]",
    "Inicio",
    "Fin",
    "Comentarios",
    "Mes1",
    "Mes2",
    "Mes3",
    "Mes4",
    "Mes5",
    "Mes6",
    "Mes7",
    "Mes8",
]

MONTH_COLUMNS = [f"Mes{i}" for i in range(1, 9)]

PO_STATUS_MAP = {
    "pendiente": "PENDING",
    "coupa generado": "COUPA_GENERATED",
    "oc enviada": "SENT",
    "enviada": "SENT",
    "aprobada": "APPROVED",
    "cerrada": "CLOSED",
    "cancelada": "CANCELLED",
}


def parse_excel_rows(file_bytes: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    missing = [col for col in REQUIRED_COLUMNS if col not in headers]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    parsed_rows: list[dict[str, Any]] = []
    for index, row in enumerate(rows[1:], start=2):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        row_data = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        row_data["_row_number"] = index
        parsed_rows.append(row_data)
    return headers, parsed_rows


def parse_decimal(value: Any) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def parse_int(value: Any) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(float(str(value)))
    except (ValueError, TypeError):
        return None


def parse_date(value: Any) -> date | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    return None


def map_po_status(value: Any) -> str:
    if value is None or str(value).strip() == "":
        return "PENDING"
    normalized = str(value).strip().lower()
    return PO_STATUS_MAP.get(normalized, "PENDING")
